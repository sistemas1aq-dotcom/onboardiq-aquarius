import io
import logging
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..models.database import get_db
from ..models.usuario import Usuario
from ..models.postulante import Postulante
from ..auth.dependencies import require_role
from ..auth.password import hash_password
from ..config import get_settings
from ..services.email_service import send_email, template_credenciales

logger = logging.getLogger(__name__)
settings = get_settings()

router = APIRouter()


@router.post("/upload")
def upload_masivo(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(require_role(["admin"])),
):
    """Carga masiva de postulantes desde archivo Excel (.xlsx)."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    try:
        contents = file.file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {str(e)}")

    # Validar cabeceras
    headers = [cell.value for cell in ws[1]] if ws.max_row and ws.max_row >= 1 else []
    expected = ["nombre", "apellido_paterno", "apellido_materno", "dni", "email", "area", "cargo", "telefono"]
    headers_lower = [str(h).strip().lower() if h else "" for h in headers]
    for col in expected:
        if col not in headers_lower:
            raise HTTPException(
                status_code=400,
                detail=f"Columna requerida '{col}' no encontrada. Cabeceras encontradas: {headers}. Descargue la plantilla actualizada.",
            )

    col_map = {col: headers_lower.index(col) for col in expected}

    creados = 0
    fallidos = 0
    errores = []
    default_password = settings.DEFAULT_PASSWORD

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            nombre_val = str(row[col_map["nombre"]] or "").strip()
            ap_paterno_val = str(row[col_map["apellido_paterno"]] or "").strip()
            ap_materno_val = str(row[col_map["apellido_materno"]] or "").strip()
            dni_val = str(row[col_map["dni"]] or "").strip()
            email_val = str(row[col_map["email"]] or "").strip()
            area_val = str(row[col_map["area"]] or "").strip()
            cargo_val = str(row[col_map["cargo"]] or "").strip()
            telefono_val = str(row[col_map["telefono"]] or "").strip()
            # Build nombre_completo from parts
            nombre_completo = " ".join(filter(None, [nombre_val, ap_paterno_val, ap_materno_val])).strip()

            # Validar campos obligatorios
            if not nombre_val:
                errores.append(f"Fila {row_num}: nombre vacío")
                fallidos += 1
                continue
            if not dni_val:
                errores.append(f"Fila {row_num}: DNI vacío")
                fallidos += 1
                continue
            if not email_val:
                errores.append(f"Fila {row_num}: email vacío")
                fallidos += 1
                continue
            if not cargo_val:
                errores.append(f"Fila {row_num}: cargo vacío")
                fallidos += 1
                continue

            # Validar duplicados
            if db.query(Usuario).filter(Usuario.dni == dni_val).first():
                errores.append(f"Fila {row_num}: DNI {dni_val} ya existe")
                fallidos += 1
                continue
            if db.query(Usuario).filter(Usuario.email == email_val).first():
                errores.append(f"Fila {row_num}: Email {email_val} ya existe")
                fallidos += 1
                continue

            # Crear usuario
            usuario = Usuario(
                email=email_val,
                password_hash=hash_password(default_password),
                nombre=nombre_completo,
                apellido_paterno=ap_paterno_val if ap_paterno_val else None,
                apellido_materno=ap_materno_val if ap_materno_val else None,
                dni=dni_val,
                rol="postulante",
                telefono=telefono_val if telefono_val else None,
                area=area_val if area_val else None,
                cargo=cargo_val if cargo_val else None,
                activo=True,
            )
            db.add(usuario)
            db.flush()

            # Crear postulante (cargo = puesto al que postula)
            postulante = Postulante(
                usuario_id=usuario.id,
                puesto=cargo_val,
                estado="En Evaluacion",
                area=area_val if area_val else None,
            )
            db.add(postulante)
            db.flush()

            # Enviar email de credenciales
            html = template_credenciales(nombre_val, email_val, default_password, cargo_val)
            send_email(email_val, f"Bienvenido a OnboardIQ Aquarius - Credenciales de acceso", html)

            creados += 1

        except Exception as e:
            logger.error(f"Error en fila {row_num}: {e}")
            errores.append(f"Fila {row_num}: {str(e)}")
            fallidos += 1
            db.rollback()
            continue

    db.commit()
    return {"creados": creados, "fallidos": fallidos, "errores": errores}


@router.get("/plantilla")
def descargar_plantilla(
    current_user: Usuario = Depends(require_role(["admin"])),
):
    """Descarga plantilla Excel para carga masiva de postulantes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Postulantes"

    headers = ["nombre", "dni", "email", "area", "cargo", "telefono"]
    header_labels = ["Nombre Completo", "DNI", "Email", "Área", "Cargo (Puesto)", "Teléfono"]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0A1F3D", end_color="0A1F3D", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Escribir cabeceras con nombres legibles
    for col_num, (header, label) in enumerate(zip(headers, header_labels), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Segunda fila con labels descriptivos
    label_font = Font(italic=True, color="64748B", size=10)
    for col_num, label in enumerate(header_labels, 1):
        cell = ws.cell(row=2, column=col_num, value=label)
        cell.font = label_font
        cell.border = thin_border

    # Filas de ejemplo
    ejemplos = [
        ["Juan Carlos Pérez García", "12345678", "jperez@email.com", "Finanzas", "Analista Financiero", "999888777"],
        ["María Elena López Ruiz", "87654321", "mlopez@email.com", "Recursos Humanos", "Analista de RRHH", "999777666"],
        ["Roberto García Quispe", "45678912", "rgarcia@email.com", "Tecnología", "Desarrollador Senior", "999666555"],
    ]
    for row_num, ejemplo in enumerate(ejemplos, 3):
        for col_num, valor in enumerate(ejemplo, 1):
            cell = ws.cell(row=row_num, column=col_num, value=valor)
            cell.border = thin_border

    # Ajustar ancho de columnas
    widths = [30, 12, 30, 20, 25, 15]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # Hoja de instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA CARGA MASIVA DE POSTULANTES"],
        [""],
        ["1. Complete los datos en la hoja 'Postulantes'"],
        ["2. La primera fila contiene los nombres de las columnas (NO modificar)"],
        ["3. La segunda fila muestra los nombres descriptivos (puede eliminarla)"],
        ["4. Las filas 3-5 son ejemplos (elimínelas antes de subir)"],
        [""],
        ["CAMPOS OBLIGATORIOS:"],
        ["  - nombre: Nombre completo del postulante"],
        ["  - dni: Documento Nacional de Identidad (8 dígitos)"],
        ["  - email: Correo electrónico (será su usuario de login)"],
        ["  - cargo: Cargo al que postula (será el puesto asignado)"],
        [""],
        ["CAMPOS OPCIONALES:"],
        ["  - area: Área de la empresa (ej: Finanzas, Tecnología, RRHH)"],
        ["  - telefono: Número de teléfono de contacto"],
        [""],
        ["NOTAS:"],
        ["  - La contraseña por defecto es: Aquarius2026"],
        ["  - Se enviará un correo automático con las credenciales"],
        ["  - Los DNI y emails deben ser únicos (no repetidos)"],
        ["  - El sistema creará automáticamente el usuario y el postulante"],
    ]
    title_font = Font(bold=True, size=14, color="0A1F3D")
    ws_inst.cell(row=1, column=1, value=instrucciones[0][0]).font = title_font
    for row_num, row in enumerate(instrucciones[1:], 2):
        ws_inst.cell(row=row_num, column=1, value=row[0] if row else "")
    ws_inst.column_dimensions["A"].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_carga_masiva_postulantes.xlsx"},
    )
